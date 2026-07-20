import json
from pathlib import Path

import pytest

from app.config import get_settings


@pytest.fixture()
def talk_script_db(tmp_path, monkeypatch):
    db_path = tmp_path / "talk_script.db"
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    get_settings.cache_clear()
    yield db_path
    get_settings.cache_clear()


def test_import_excel_loads_valid_talk_script_library(talk_script_db, tmp_path):
    from openpyxl import Workbook

    from app.talk_script.excel_importer import import_talk_script_excel
    from app.talk_script.repository import get_active_template_by_question_id

    workbook_path = tmp_path / "scripts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "scene_index"
    ws.append(
        [
            "scene_id",
            "scene_name",
            "scene_definition",
            "enter_conditions",
            "typical_user_messages",
            "exclude_conditions",
            "priority",
            "status",
        ]
    )
    ws.append(["S04", "养护固定问答库", "", "", "换盆", "", 60, "active"])
    ws = wb.create_sheet("question_cluster")
    ws.append(
        [
            "question_id",
            "scene_id",
            "sub_scene_name",
            "standard_question",
            "core_intent",
            "user_question_aliases",
            "positive_examples",
            "negative_examples",
            "keywords",
            "required_conditions",
            "exclude_conditions",
            "default_template_id",
            "confidence_threshold",
            "priority",
            "status",
        ]
    )
    ws.append(
        [
            "Q04_01_001",
            "S04",
            "收到后换盆",
            "兰花收到后是否需要换盆？",
            "询问收到后换盆",
            "收到要换盆吗｜刚收到要不要换盆",
            "刚收到兰花要不要换盆",
            "烂根了要不要换盆",
            "收到｜换盆",
            "",
            "烂根",
            "T04_01_001",
            0.75,
            80,
            "active",
        ]
    )
    ws = wb.create_sheet("template_library")
    ws.append(
        [
            "template_id",
            "question_id",
            "template_name",
            "answer_default",
            "answer_goal",
            "need_slot_filling",
            "handoff_rule",
            "status",
            "version",
            "change_note",
        ]
    )
    ws.append(
        [
            "T04_01_001",
            "Q04_01_001",
            "收到后换盆",
            "收到后先不要急着换盆。",
            "",
            "no",
            "",
            "active",
            "v1",
            "",
        ]
    )
    ws = wb.create_sheet("field_dictionary")
    ws.append(["table_name", "field_name", "field_meaning", "required", "example", "note"])
    wb.save(workbook_path)

    result = import_talk_script_excel(workbook_path)

    assert result == {"scene_count": 1, "question_count": 1, "template_count": 1}
    template = get_active_template_by_question_id("Q04_01_001")
    assert template is not None
    assert template.answer_default == "收到后先不要急着换盆。"


def test_first_order_seed_is_repeatable_and_structured(talk_script_db):
    from app.talk_script.first_order_sales_seed import ensure_first_order_sales_templates
    from app.talk_script.repository import list_sales_templates

    assert ensure_first_order_sales_templates() > 0
    assert ensure_first_order_sales_templates() == 0
    rows = list_sales_templates(sales_stage="closing")
    assert rows
    assert all(row.sales_action and row.branch_code for row in rows)
    assert any(json.loads(row.required_fact_keys_json) for row in rows)


def _seed_single_talk_script_question():
    from app.talk_script.repository import replace_talk_script_library

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "care",
                "typical_user_messages": "care orchid root leaf fertilizer refund",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "care question",
                "keywords": "care root leaf fertilizer refund",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.75,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "answer_default": "care answer",
                "status": "active",
            }
        ],
    )


@pytest.mark.asyncio
async def test_match_talk_script_passes_through_when_classifier_needs_slot_filling(
    talk_script_db,
):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.service import match_talk_script

    _seed_single_talk_script_question()

    async def classifier(**kwargs):
        del kwargs
        return ClassifierDecision(
            matched=False,
            question_id=None,
            confidence=0.0,
            need_slot_filling=True,
            need_human=False,
            reason="missing key care details",
        )

    result = await match_talk_script(
        customer_id="cust_slot",
        current_message="care fertilizer",
        classifier=classifier,
    )

    assert result.status == "pass_through"
    assert result.need_human is False
    assert result.reason == "need_slot_filling"


@pytest.mark.asyncio
async def test_match_talk_script_passes_through_when_classifier_marks_care_issue_human(
    talk_script_db,
):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.service import match_talk_script

    _seed_single_talk_script_question()

    async def classifier(**kwargs):
        del kwargs
        return ClassifierDecision(
            matched=False,
            question_id=None,
            confidence=0.9,
            need_slot_filling=False,
            need_human=True,
            reason="severe orchid care issue",
        )

    result = await match_talk_script(
        customer_id="cust_care_handoff",
        current_message="care root leaf problem",
        classifier=classifier,
    )

    assert result.status == "pass_through"
    assert result.need_human is False
    assert result.reason == "need_human_non_critical"


@pytest.mark.asyncio
async def test_match_talk_script_keeps_handoff_for_refund_request(talk_script_db):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.service import match_talk_script

    _seed_single_talk_script_question()

    async def classifier(**kwargs):
        del kwargs
        return ClassifierDecision(
            matched=False,
            question_id=None,
            confidence=0.95,
            need_slot_filling=False,
            need_human=True,
            reason="refund requested",
        )

    result = await match_talk_script(
        customer_id="cust_refund",
        current_message="refund and human support",
        classifier=classifier,
    )

    assert result.status == "handoff"
    assert result.need_human is True
    assert result.reason == "need_human"


@pytest.mark.asyncio
async def test_match_talk_script_routes_common_price_objection_to_trade_scene(
    talk_script_db,
):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.repository import replace_talk_script_library
    from app.talk_script.service import match_talk_script

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S07",
                "scene_name": "交易异议与成交转化库",
                "typical_user_messages": "多少钱｜怎么卖｜太贵了｜能不能便宜点",
                "priority": 80,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q07_01_001",
                "scene_id": "S07",
                "standard_question": "用户觉得价格贵",
                "keywords": "贵｜太贵｜有点贵｜便宜",
                "default_template_id": "T07_01_001",
                "confidence_threshold": 0.75,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T07_01_001",
                "question_id": "Q07_01_001",
                "answer_default": "我理解你会关注价格。",
                "status": "active",
            }
        ],
    )

    async def classifier(**kwargs):
        assert [item["question_id"] for item in kwargs["candidate_questions"]] == [
            "Q07_01_001"
        ]
        return ClassifierDecision(
            matched=True,
            question_id="Q07_01_001",
            confidence=0.9,
            reason="命中价格异议",
        )

    result = await match_talk_script(
        customer_id="user_001",
        current_message="这个有点贵",
        trace_id="trace_price_001",
        classifier=classifier,
    )

    assert result.status == "matched"
    assert result.scene_id == "S07"
    assert result.question_id == "Q07_01_001"
    assert result.template_id == "T07_01_001"
    assert result.answer == "我理解你会关注价格。"


@pytest.mark.asyncio
async def test_match_talk_script_returns_fixed_answer_for_high_confidence(talk_script_db):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.repository import replace_talk_script_library
    from app.talk_script.service import match_talk_script

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "养护固定问答库",
                "typical_user_messages": "收到要换盆吗｜怎么换盆",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "兰花收到后是否需要换盆？",
                "core_intent": "询问收到后换盆",
                "user_question_aliases": "收到要换盆吗｜刚收到要不要换盆",
                "positive_examples": "刚收到兰花要不要换盆",
                "negative_examples": "我的兰花烂根了",
                "keywords": "收到｜换盆",
                "exclude_conditions": "烂根",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.75,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "template_name": "收到后换盆",
                "answer_default": "收到后先不要急着换盆。",
                "status": "active",
            }
        ],
    )

    async def classifier(**kwargs):
        assert [item["question_id"] for item in kwargs["candidate_questions"]] == [
            "Q04_01_001"
        ]
        return ClassifierDecision(
            matched=True,
            question_id="Q04_01_001",
            confidence=0.92,
            reason="候选中明确命中收到后换盆",
        )

    result = await match_talk_script(
        customer_id="user_001",
        current_message="刚收到兰花要不要换盆？",
        trace_id="trace_001",
        classifier=classifier,
    )

    assert result.status == "matched"
    assert result.answer == "收到后先不要急着换盆。"
    assert result.scene_id == "S04"
    assert result.question_id == "Q04_01_001"
    assert result.template_id == "T04_01_001"


@pytest.mark.asyncio
async def test_match_talk_script_passes_through_when_template_already_sent_to_customer(
    talk_script_db,
):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.repository import replace_talk_script_library
    from app.talk_script.service import match_talk_script

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "care",
                "typical_user_messages": "again",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "repeat care script",
                "keywords": "again",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.75,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "answer_default": "Fixed script once.",
                "status": "active",
            }
        ],
    )

    async def classifier(**kwargs):
        return ClassifierDecision(
            matched=True,
            question_id="Q04_01_001",
            confidence=0.95,
            reason="matched",
        )

    first = await match_talk_script(
        customer_id="user_001",
        current_message="again",
        trace_id="trace_once_1",
        classifier=classifier,
    )
    second = await match_talk_script(
        customer_id="user_001",
        current_message="again",
        trace_id="trace_once_2",
        classifier=classifier,
    )

    assert first.status == "matched"
    assert first.answer == "Fixed script once."
    assert second.status == "pass_through"
    assert second.reason == "template_already_sent"
    assert second.answer == ""


@pytest.mark.asyncio
async def test_match_talk_script_passes_through_when_confidence_is_low(talk_script_db):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.repository import replace_talk_script_library
    from app.talk_script.service import match_talk_script

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "养护固定问答库",
                "typical_user_messages": "换盆",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "兰花收到后是否需要换盆？",
                "keywords": "换盆",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.8,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "answer_default": "收到后先不要急着换盆。",
                "status": "active",
            }
        ],
    )

    async def classifier(**kwargs):
        return ClassifierDecision(
            matched=True,
            question_id="Q04_01_001",
            confidence=0.5,
            reason="不够确定",
        )

    result = await match_talk_script(
        customer_id="user_001",
        current_message="这个要不要换盆？",
        trace_id="trace_002",
        classifier=classifier,
    )

    assert result.status == "pass_through"
    assert result.need_human is False
    assert result.answer == ""
    assert result.reason == "confidence_below_threshold"


@pytest.mark.asyncio
async def test_match_talk_script_passes_through_when_no_scene_matches(talk_script_db):
    from app.talk_script.service import match_talk_script

    result = await match_talk_script(
        customer_id="user_001",
        current_message="帮我总结一下这篇资料",
        trace_id="trace_003",
    )

    assert result.status == "pass_through"
    assert result.answer == ""
    assert result.need_human is False


@pytest.mark.asyncio
async def test_chat_orchestrator_skips_talk_script_for_rag_route(
    talk_script_db, monkeypatch
):
    from app.schemas.chat import ChatRequest
    from app.schemas.intent import IntentResult
    from app.services import chat_orchestrator
    from app.talk_script.repository import replace_talk_script_library

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "养护固定问答库",
                "typical_user_messages": "换盆",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "兰花收到后是否需要换盆？",
                "keywords": "收到｜换盆",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.5,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "answer_default": "收到后先不要急着换盆。",
                "status": "active",
            }
        ],
    )

    async def fake_retrieve_intent_examples(message, top_k):
        del message, top_k
        return []

    async def fake_classify_intent(message, user_state, candidates):
        del message, user_state, candidates
        return IntentResult(
            route="rag_answer",
            primary_intent="care_question",
            confidence=0.9,
            need_rag=True,
        )

    async def fake_rag(message, user_state, policy_decision=None):
        del message, user_state, policy_decision
        return {"answer": "收到后先不要急着换盆。", "sources": []}

    monkeypatch.setattr(
        chat_orchestrator, "retrieve_intent_examples", fake_retrieve_intent_examples
    )
    monkeypatch.setattr(chat_orchestrator, "classify_intent", fake_classify_intent)
    monkeypatch.setattr("app.services.rag_service.answer_knowledge", fake_rag)

    result = await chat_orchestrator.handle_chat(
        ChatRequest(
            channel="api",
            user_id="user_001",
            message="刚收到兰花要不要换盆？",
            kb_id="kb_default",
        )
    )

    assert result["answer"] == "收到后先不要急着换盆。"
    assert result["reply_type"] == "rag"
    assert result["route"] == "rag_answer"
    assert result["template"] == {}
    assert result["need_human"] is False


@pytest.mark.asyncio
async def test_chat_orchestrator_hydrates_profile_before_policy_and_reply(
    talk_script_db, monkeypatch
):
    from app.schemas.chat import ChatRequest
    from app.schemas.intent import IntentResult
    from app.services import chat_orchestrator

    captured = {}

    async def fake_get_profile_bundle(user_id):
        assert user_id == "user_profiled"
        return {
            "profile": {
                "customer_tags": ["浙江省", "100-200盆", "L3 黄金期", "建兰"],
                "ai_summary": "客户在浙江，养了100盆花，正在咨询推荐。",
                "pain_points": ["想基于浙江环境推荐建兰"],
            },
            "recent_memories": [
                {"role": "customer", "content": "那我要买，给我链接"},
                {"role": "assistant", "content": "好的收到了解"},
            ],
            "events": [],
        }

    async def fake_retrieve_intent_examples(message, top_k):
        del message, top_k
        return []

    async def fake_classify_intent(message, user_state, candidates):
        del message, candidates
        captured["state_tags_at_intent"] = list(user_state.customer_tags)
        return IntentResult(
            route="rag_answer",
            primary_intent="knowledge_question",
            confidence=0.9,
            need_rag=True,
        )

    async def fake_answer_knowledge(message, user_state, policy_decision=None):
        del message, policy_decision
        captured["metadata_profile"] = user_state.metadata.get("profile")
        captured["recent_turns"] = user_state.metadata.get("recent_turns")
        return {"answer": "基于浙江和建兰推荐。", "sources": []}

    monkeypatch.setattr(chat_orchestrator, "get_profile_bundle", fake_get_profile_bundle)
    monkeypatch.setattr(
        chat_orchestrator, "retrieve_intent_examples", fake_retrieve_intent_examples
    )
    monkeypatch.setattr(chat_orchestrator, "classify_intent", fake_classify_intent)
    monkeypatch.setattr("app.services.rag_service.answer_knowledge", fake_answer_knowledge)

    result = await chat_orchestrator.handle_chat(
        ChatRequest(
            channel="api",
            user_id="user_profiled",
            message="你们家建兰有什么推荐",
            kb_id="kb_default",
        )
    )

    assert result["answer"] == "基于浙江和建兰推荐。"
    assert captured["state_tags_at_intent"] == [
        "浙江省",
        "100-200盆",
        "L3 黄金期",
        "建兰",
    ]
    assert captured["metadata_profile"]["ai_summary"] == "客户在浙江，养了100盆花，正在咨询推荐。"
    assert captured["metadata_profile"]["pain_points"] == ["想基于浙江环境推荐建兰"]
    assert captured["recent_turns"][0]["content"] == "那我要买，给我链接"


def test_to_chat_data_merges_short_blank_line_paragraphs():
    from app.schemas.intent import IntentResult
    from app.schemas.reply import FinalReply
    from app.services.chat_orchestrator import _to_chat_data

    data = _to_chat_data(
        "session_1",
        "trace_1",
        IntentResult(
            route="rag_answer",
            primary_intent="knowledge_question",
            confidence=0.9,
        ),
        FinalReply(
            answer="第一段。\n\n第二段。\n第三段。",
            reply_type="rag",
            route="rag_answer",
        ),
    )

    assert data["answer"] == "第一段。\n第二段。\n第三段。"
    assert data["answer_segments"] == ["第一段。第二段。 第三段。"]


def test_to_chat_data_merges_short_structured_answer_segments():
    from app.schemas.intent import IntentResult
    from app.schemas.reply import FinalReply
    from app.services.chat_orchestrator import _to_chat_data

    data = _to_chat_data(
        "session_1",
        "trace_1",
        IntentResult(
            route="template_reply",
            primary_intent="ask_price",
            confidence=0.9,
        ),
        FinalReply(
            answer="完整回答。您有多少盆？",
            answer_segments=["完整回答。", "您有多少盆？"],
            reply_type="template",
            route="template_reply",
        ),
    )

    assert data["answer_segments"] == ["完整回答。您有多少盆？"]


def test_answer_segments_keeps_short_reply_as_one_message():
    from app.services.chat_orchestrator import _answer_segments

    sentences = [f"第{index}句。" for index in range(1, 6)]
    answer = "".join(sentences)

    segments = _answer_segments(answer)

    assert segments == [answer]


def test_answer_segments_removes_markdown_without_forcing_short_split():
    from app.services.chat_orchestrator import _answer_segments

    answer = (
        "1. **脱盆修根**：剪掉烂根，用多菌灵等杀菌剂消毒，"
        "然后放在通风处晾干。"
    )

    segments = _answer_segments(answer)

    assert "**" not in "".join(segments)
    assert not segments[0].startswith("1.")
    assert len(segments) == 1


def test_plain_customer_text_removes_markdown_tables_and_keeps_fenced_content():
    from app.services.customer_reply_formatter import plain_customer_text

    text = """# 处理建议
- **修根**
---
~~别急~~
| 步骤 | 处理 |
| --- | --- |
| 一 | 晾根 |
```
保持通风
```
"""

    assert plain_customer_text(text) == "处理建议\n修根\n别急\n步骤 处理\n一 晾根\n保持通风"


def test_ordinary_logistics_question_is_not_a_critical_handoff():
    from app.talk_script.service import _is_critical_human_request

    assert not _is_critical_human_request("物流什么时候到", "need_human")
