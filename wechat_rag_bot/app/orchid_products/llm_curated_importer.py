from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.orchid_products.excel_importer import OrchidImportPayload


VARIETY_SHEETS = {"建兰", "春兰", "春剑", "寒兰", "墨兰", "莲瓣兰", "蕙兰", "其他品类"}
COMMON_KNOWLEDGE_SHEET = "通用知识点"
SKU_SHEET = "链接详情"
SALES_COPY_SHEET = "私域运营团队2"
HOT_BREAKDOWN_SHEET = "热门品种拆解"


def build_llm_curated_payload(path: str | Path) -> OrchidImportPayload:
    workbook = load_workbook(path, read_only=True, data_only=True)
    payload = OrchidImportPayload(
        categories=[],
        varieties=[],
        traits=[],
        value_points=[],
        skus=[],
        common_knowledge=[],
        sales_copy=[],
        hot_breakdowns=[],
        knowledge_chunks=[],
    )
    seen_categories: set[str] = set()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name in VARIETY_SHEETS:
            if sheet_name not in seen_categories:
                payload.categories.append(
                    {"category_name": sheet_name, "category_description": None}
                )
                seen_categories.add(sheet_name)
            _curate_variety_rows(payload, sheet_name, _read_records(sheet, ["品种名"]))
        elif sheet_name == COMMON_KNOWLEDGE_SHEET:
            _curate_common_rows(payload, _read_records(sheet, ["分类", "共享话术"]))
        elif sheet_name == SKU_SHEET:
            _curate_sku_rows(payload, _read_records(sheet, ["品种", "品名"]))
        elif sheet_name == SALES_COPY_SHEET:
            _curate_sales_rows(payload, _read_records(sheet, ["填写人", "品种名"]))
        elif sheet_name == HOT_BREAKDOWN_SHEET:
            _curate_hot_rows(payload, _read_records(sheet, ["品种名", "品类"]))

    return payload


def _curate_variety_rows(
    payload: OrchidImportPayload,
    category_name: str,
    records: list[dict[str, Any]],
) -> None:
    for record in records:
        variety_name = _text(record.get("品种名"))
        if not variety_name:
            continue
        alias = _text(record.get("别名")) or None
        raw_basic_info = _text(record.get("产品基础信息")) or None
        payload.varieties.append(
            {
                "category_id": None,
                "category_name": category_name,
                "variety_name": variety_name,
                "primary_alias": alias,
                "aliases_text": alias,
                "source_type": None,
                "origin_area": None,
                "history_background": None,
                "summary": None,
                "suitable_level": _text(record.get("匹配人群")) or None,
                "base_spec": _text(record.get("基础规格")) or None,
                "base_price_text": _text(record.get("价格")) or None,
                "raw_basic_info": raw_basic_info,
            }
        )
        if raw_basic_info:
            payload.value_points.append(
                {
                    "variety_id": None,
                    "variety_name": variety_name,
                    "category_name": category_name,
                    "value_type": "产品基础信息",
                    "title": f"{variety_name}产品基础信息",
                    "content": raw_basic_info,
                    "keywords": None,
                }
            )
            payload.knowledge_chunks.append(
                _chunk(
                    "orchid_varieties",
                    "variety",
                    variety_name,
                    category_name,
                    "产品基础信息",
                    f"{variety_name} - 产品基础信息",
                    raw_basic_info,
                )
            )


def _curate_common_rows(
    payload: OrchidImportPayload,
    records: list[dict[str, Any]],
) -> None:
    for record in records:
        category = _text(record.get("分类"))
        content = _text(record.get("共享话术"))
        if not category or not content:
            continue
        row = {
            "knowledge_category": category,
            "knowledge_type": _common_knowledge_type(category, content),
            "applies_to_category": _category_in_text(category),
            "content": content,
        }
        payload.common_knowledge.append(row)
        payload.knowledge_chunks.append(
            _chunk(
                "orchid_common_knowledge",
                "common_knowledge",
                None,
                row["applies_to_category"],
                "通用知识",
                category,
                content,
            )
        )


def _curate_sku_rows(
    payload: OrchidImportPayload,
    records: list[dict[str, Any]],
) -> None:
    previous_name = ""
    for record in records:
        variety_name = _text(record.get("品名")) or previous_name
        if not variety_name:
            continue
        previous_name = variety_name
        price_text = _text(record.get("到手价"))
        payload.skus.append(
            {
                "category_name": _text(record.get("品种")) or None,
                "variety_name": variety_name,
                "seedling_count": _text(record.get("苗数")) or None,
                "package_spec": _text(record.get("规格")) or None,
                "flower_bud_status": _text(record.get("有无花苞")) or None,
                "price": _price(price_text),
                "price_text": price_text or None,
            }
        )


def _curate_sales_rows(
    payload: OrchidImportPayload,
    records: list[dict[str, Any]],
) -> None:
    chunk_fields = [
        ("适合人群", "适合人群"),
        ("产品背景", "产品背景"),
        ("叶姿", "叶姿"),
        ("瓣型", "瓣型"),
        ("花色", "花色"),
        ("香味", "香味"),
        ("花期", "花期"),
        ("养护难度", "养护难度"),
        ("使用场景", "使用场景"),
        ("卖点", "卖点"),
    ]
    for record in records:
        variety_name = _text(record.get("品种名"))
        if not variety_name:
            continue
        row = {
            "writer_name": _text(record.get("填写人")) or None,
            "variety_name": variety_name,
            "target_audience": _text(record.get("适合人群")) or None,
            "product_background": _text(record.get("产品背景")) or None,
            "leaf_posture": _text(record.get("叶姿")) or None,
            "petal_type": _text(record.get("瓣型")) or None,
            "flower_color": _text(record.get("花色")) or None,
            "fragrance": _text(record.get("香味")) or None,
            "flowering_period": _text(record.get("花期")) or None,
            "care_difficulty": _text(record.get("养护难度")) or None,
            "usage_scene": _text(record.get("使用场景")) or None,
            "selling_points": _text(record.get("卖点")) or None,
        }
        payload.sales_copy.append(row)
        for source_field, chunk_type in chunk_fields:
            content = _text(record.get(source_field))
            if content:
                payload.knowledge_chunks.append(
                    _chunk(
                        "orchid_sales_copy",
                        "sales_copy",
                        variety_name,
                        None,
                        chunk_type,
                        f"{variety_name} - {chunk_type}话术",
                        content,
                    )
                )


def _curate_hot_rows(
    payload: OrchidImportPayload,
    records: list[dict[str, Any]],
) -> None:
    for record in records:
        variety_name = _first_line(record.get("品种名"))
        if not variety_name:
            continue
        category_name = _text(record.get("品类")) or None
        row = {
            "variety_name": variety_name,
            "category_name": category_name,
            "status_history_supply_price_authenticity": _first_value_with_prefix(
                record, "品种地位"
            ),
            "aesthetic_traits": _first_value_with_prefix(record, "瓣"),
            "cultivation_care": _first_value_with_prefix(record, "栽培"),
            "consensus_reputation": _first_value_with_prefix(record, "共识"),
        }
        row["raw_text"] = "\n".join(
            value for value in row.values() if isinstance(value, str) and value
        )
        payload.hot_breakdowns.append(row)
        if row["aesthetic_traits"]:
            payload.traits.append(
                {
                    "variety_id": None,
                    "variety_name": variety_name,
                    "category_name": category_name,
                    "trait_type": "综合审美特征",
                    "trait_value": row["aesthetic_traits"],
                    "keywords": None,
                }
            )
        for field, chunk_type in [
            ("status_history_supply_price_authenticity", "品种地位与交易认知"),
            ("aesthetic_traits", "瓣色香姿神"),
            ("cultivation_care", "栽培与养护"),
            ("consensus_reputation", "共识口碑"),
        ]:
            content = row.get(field)
            if content:
                payload.knowledge_chunks.append(
                    _chunk(
                        "orchid_hot_breakdowns",
                        "hot_breakdown",
                        variety_name,
                        category_name,
                        chunk_type,
                        f"{variety_name} - {chunk_type}",
                        content,
                    )
                )


def _read_records(sheet, required_headers: list[str]) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = None
    headers: list[str] = []
    for index, row in enumerate(rows):
        candidate = [_text(cell) for cell in row]
        if all(header in candidate for header in required_headers):
            header_index = index
            headers = candidate
            break
    if header_index is None:
        return []

    records = []
    for row in rows[header_index + 1 :]:
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        if any(_text(value) for value in record.values()):
            records.append(record)
    return records


def _chunk(
    source_table: str,
    entity_type: str,
    variety_name: str | None,
    category_name: str | None,
    chunk_type: str,
    chunk_title: str,
    content: str,
) -> dict[str, Any]:
    return {
        "source_table": source_table,
        "source_id": None,
        "entity_type": entity_type,
        "variety_name": variety_name,
        "category_name": category_name,
        "chunk_type": chunk_type,
        "chunk_title": chunk_title,
        "content": content,
        "embedding_json": None,
    }


def _common_knowledge_type(category: str, content: str) -> str:
    text = f"{category}\n{content}"
    if "种源" in text or "解释" in text:
        return "术语解释"
    if "劣势" in text:
        return "品类劣势"
    if "优势" in text:
        return "品类优势"
    if "养" in text or "浇水" in text or "施肥" in text:
        return "养护知识"
    if "文化" in text or "寓意" in text:
        return "文化价值"
    return "销售话术"


def _category_in_text(text: str) -> str | None:
    for category in VARIETY_SHEETS:
        if category != "其他品类" and category in text:
            return category
    return None


def _first_value_with_prefix(record: dict[str, Any], prefix: str) -> str | None:
    for key, value in record.items():
        if str(key).startswith(prefix):
            return _text(value) or None
    return None


def _first_line(value: Any) -> str:
    return _text(value).split("\n", 1)[0].strip()


def _price(value: str) -> float | None:
    digits = []
    seen_dot = False
    for char in value:
        if char.isdigit():
            digits.append(char)
        elif char == "." and digits and not seen_dot:
            digits.append(char)
            seen_dot = True
        elif digits:
            break
    if not digits:
        return None
    return float("".join(digits).rstrip("."))


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").strip()
