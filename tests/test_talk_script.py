from pathlib import Path

import pytest

from app.config import get_settings


@pytest.fixture()
def talk_script_db(tmp_path, monkeypatch):
    db_path = tmp_path / "talk_script.db"
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    get_settings.cache_clear()
    yield db_path
    get_settings.cache_clear()


def test_import_excel_loads_valid_talk_script_library(talk_script_db, tmp_path):
    from openpyxl import Workbook

    from app.talk_script.excel_importer import import_talk_script_excel
    from app.talk_script.repository import get_active_template_by_question_id

    workbook_path = tmp_path / "scripts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "scene_index"
    ws.append(
        [
            "scene_id",
            "scene_name",
            "scene_definition",
            "enter_conditions",
            "typical_user_messages",
            "exclude_conditions",
            "priority",
            "status",
        ]
    )
    ws.append(["S04", "养护固定问答库", "", "", "换盆", "", 60, "active"])
    ws = wb.create_sheet("question_cluster")
    ws.append(
        [
            "question_id",
            "scene_id",
            "sub_scene_name",
            "standard_question",
            "core_intent",
            "user_question_aliases",
            "positive_examples",
            "negative_examples",
            "keywords",
            "required_conditions",
            "exclude_conditions",
            "default_template_id",
            "confidence_threshold",
            "priority",
            "status",
        ]
    )
    ws.append(
        [
            "Q04_01_001",
            "S04",
            "收到后换盆",
            "兰花收到后是否需要换盆？",
            "询问收到后换盆",
            "收到要换盆吗｜刚收到要不要换盆",
            "刚收到兰花要不要换盆",
            "烂根了要不要换盆",
            "收到｜换盆",
            "",
            "烂根",
            "T04_01_001",
            0.75,
            80,
            "active",
        ]
    )
    ws = wb.create_sheet("template_library")
    ws.append(
        [
            "template_id",
            "question_id",
            "template_name",
            "answer_default",
            "answer_goal",
            "need_slot_filling",
            "handoff_rule",
            "status",
            "version",
            "change_note",
        ]
    )
    ws.append(
        [
            "T04_01_001",
            "Q04_01_001",
            "收到后换盆",
            "收到后先不要急着换盆。",
            "",
            "no",
            "",
            "active",
            "v1",
            "",
        ]
    )
    ws = wb.create_sheet("field_dictionary")
    ws.append(["table_name", "field_name", "field_meaning", "required", "example", "note"])
    wb.save(workbook_path)

    result = import_talk_script_excel(workbook_path)

    assert result == {"scene_count": 1, "question_count": 1, "template_count": 1}
    template = get_active_template_by_question_id("Q04_01_001")
    assert template is not None
    assert template.answer_default == "收到后先不要急着换盆。"


@pytest.mark.asyncio
async def test_match_talk_script_returns_fixed_answer_for_high_confidence(talk_script_db):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.repository import replace_talk_script_library
    from app.talk_script.service import match_talk_script

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "养护固定问答库",
                "typical_user_messages": "收到要换盆吗｜怎么换盆",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "兰花收到后是否需要换盆？",
                "core_intent": "询问收到后换盆",
                "user_question_aliases": "收到要换盆吗｜刚收到要不要换盆",
                "positive_examples": "刚收到兰花要不要换盆",
                "negative_examples": "我的兰花烂根了",
                "keywords": "收到｜换盆",
                "exclude_conditions": "烂根",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.75,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "template_name": "收到后换盆",
                "answer_default": "收到后先不要急着换盆。",
                "status": "active",
            }
        ],
    )

    async def classifier(**kwargs):
        assert [item["question_id"] for item in kwargs["candidate_questions"]] == [
            "Q04_01_001"
        ]
        return ClassifierDecision(
            matched=True,
            question_id="Q04_01_001",
            confidence=0.92,
            reason="候选中明确命中收到后换盆",
        )

    result = await match_talk_script(
        customer_id="user_001",
        current_message="刚收到兰花要不要换盆？",
        trace_id="trace_001",
        classifier=classifier,
    )

    assert result.status == "matched"
    assert result.answer == "收到后先不要急着换盆。"
    assert result.scene_id == "S04"
    assert result.question_id == "Q04_01_001"
    assert result.template_id == "T04_01_001"


@pytest.mark.asyncio
async def test_match_talk_script_handoffs_when_confidence_is_low(talk_script_db):
    from app.talk_script.models import ClassifierDecision
    from app.talk_script.repository import replace_talk_script_library
    from app.talk_script.service import match_talk_script

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "养护固定问答库",
                "typical_user_messages": "换盆",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "兰花收到后是否需要换盆？",
                "keywords": "换盆",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.8,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "answer_default": "收到后先不要急着换盆。",
                "status": "active",
            }
        ],
    )

    async def classifier(**kwargs):
        return ClassifierDecision(
            matched=True,
            question_id="Q04_01_001",
            confidence=0.5,
            reason="不够确定",
        )

    result = await match_talk_script(
        customer_id="user_001",
        current_message="这个要不要换盆？",
        trace_id="trace_002",
        classifier=classifier,
    )

    assert result.status == "handoff"
    assert result.need_human is True
    assert result.answer == ""
    assert result.reason == "confidence_below_threshold"


@pytest.mark.asyncio
async def test_match_talk_script_passes_through_when_no_scene_matches(talk_script_db):
    from app.talk_script.service import match_talk_script

    result = await match_talk_script(
        customer_id="user_001",
        current_message="帮我总结一下这篇资料",
        trace_id="trace_003",
    )

    assert result.status == "pass_through"
    assert result.answer == ""
    assert result.need_human is False


@pytest.mark.asyncio
async def test_chat_orchestrator_uses_talk_script_before_rag(talk_script_db, monkeypatch):
    from app.schemas.chat import ChatRequest
    from app.schemas.intent import IntentResult
    from app.services import chat_orchestrator
    from app.talk_script.repository import replace_talk_script_library

    replace_talk_script_library(
        scenes=[
            {
                "scene_id": "S04",
                "scene_name": "养护固定问答库",
                "typical_user_messages": "换盆",
                "priority": 60,
                "status": "active",
            }
        ],
        questions=[
            {
                "question_id": "Q04_01_001",
                "scene_id": "S04",
                "standard_question": "兰花收到后是否需要换盆？",
                "keywords": "收到｜换盆",
                "default_template_id": "T04_01_001",
                "confidence_threshold": 0.5,
                "priority": 80,
                "status": "active",
            }
        ],
        templates=[
            {
                "template_id": "T04_01_001",
                "question_id": "Q04_01_001",
                "answer_default": "收到后先不要急着换盆。",
                "status": "active",
            }
        ],
    )

    async def fake_retrieve_intent_examples(message, top_k):
        del message, top_k
        return []

    async def fake_classify_intent(message, user_state, candidates):
        del message, user_state, candidates
        return IntentResult(
            route="rag_answer",
            primary_intent="care_question",
            confidence=0.9,
            need_rag=True,
        )

    async def fail_rag(message, user_state):
        del message, user_state
        raise AssertionError("RAG should not run when talk script matches")

    monkeypatch.setattr(
        chat_orchestrator, "retrieve_intent_examples", fake_retrieve_intent_examples
    )
    monkeypatch.setattr(chat_orchestrator, "classify_intent", fake_classify_intent)
    monkeypatch.setattr("app.services.rag_service.answer_knowledge", fail_rag)

    result = await chat_orchestrator.handle_chat(
        ChatRequest(
            channel="api",
            user_id="user_001",
            message="刚收到兰花要不要换盆？",
            kb_id="kb_default",
        )
    )

    assert result["answer"] == "收到后先不要急着换盆。"
    assert result["reply_type"] == "template"
    assert result["route"] == "template_reply"
    assert result["template"] == {"template_id": "T04_01_001"}
    assert result["need_human"] is False
