from dataclasses import dataclass
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


VARIETY_SHEETS = {"建兰", "春兰", "春剑", "寒兰", "墨兰", "莲瓣兰", "蕙兰", "其他品类"}
COMMON_KNOWLEDGE_SHEET = "通用知识点"
SKU_SHEET = "链接详情"
SALES_COPY_SHEET = "私域运营团队2"
HOT_BREAKDOWN_SHEET = "热门品种拆解"

SECTION_LABELS = {
    "产品来源": "产品来源",
    "来源": "产品来源",
    "故事背景": "故事背景",
    "产品背景": "故事背景",
    "产品特征": "产品特征",
    "特征": "产品特征",
    "差异性": "差异性",
    "市场价": "市场价",
    "使用场景": "使用场景",
    "摆放场地": "使用场景",
    "总结": "总结",
}
TRAIT_LABELS = {
    "花色",
    "瓣型",
    "花型",
    "叶姿",
    "叶艺",
    "香味",
    "香气",
    "花期",
    "花量",
    "花守",
    "株型",
    "花杆",
    "花舌",
    "养护难度",
    "春化要求",
    "抗性",
    "生长特性",
    "适合场景",
    "适合人群",
    "寓意",
    "收藏价值",
    "真假鉴别",
}
VALUE_LABELS = {
    "产品来源",
    "故事背景",
    "差异性",
    "市场价",
    "使用场景",
    "总结",
}


@dataclass
class OrchidImportPayload:
    categories: list[dict[str, Any]]
    varieties: list[dict[str, Any]]
    traits: list[dict[str, Any]]
    value_points: list[dict[str, Any]]
    skus: list[dict[str, Any]]
    common_knowledge: list[dict[str, Any]]
    sales_copy: list[dict[str, Any]]
    hot_breakdowns: list[dict[str, Any]]
    knowledge_chunks: list[dict[str, Any]]

    @property
    def counts(self) -> dict[str, int]:
        return {
            "categories": len(self.categories),
            "varieties": len(self.varieties),
            "traits": len(self.traits),
            "value_points": len(self.value_points),
            "skus": len(self.skus),
            "common_knowledge": len(self.common_knowledge),
            "sales_copy": len(self.sales_copy),
            "hot_breakdowns": len(self.hot_breakdowns),
            "knowledge_chunks": len(self.knowledge_chunks),
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "counts": self.counts,
            "categories": self.categories,
            "varieties": self.varieties,
            "traits": self.traits,
            "value_points": self.value_points,
            "skus": self.skus,
            "common_knowledge": self.common_knowledge,
            "sales_copy": self.sales_copy,
            "hot_breakdowns": self.hot_breakdowns,
            "knowledge_chunks": self.knowledge_chunks,
        }


def build_import_payload(path: str | Path) -> OrchidImportPayload:
    workbook = load_workbook(path, read_only=True, data_only=True)
    categories: list[dict[str, Any]] = []
    varieties: list[dict[str, Any]] = []
    traits: list[dict[str, Any]] = []
    value_points: list[dict[str, Any]] = []
    skus: list[dict[str, Any]] = []
    common_knowledge: list[dict[str, Any]] = []
    sales_copy: list[dict[str, Any]] = []
    hot_breakdowns: list[dict[str, Any]] = []
    chunks: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name in VARIETY_SHEETS:
            if not any(row["category_name"] == sheet_name for row in categories):
                categories.append({"category_name": sheet_name, "category_description": None})
            records = _read_records(sheet, ["品种名"])
            for record in records:
                variety_name = _text(record.get("品种名"))
                if not variety_name:
                    continue
                raw_basic_info = _text(record.get("产品基础信息"))
                extracted = _extract_basic_info(raw_basic_info)
                variety = {
                    "category_id": None,
                    "category_name": sheet_name,
                    "variety_name": variety_name,
                    "primary_alias": _text(record.get("别名")) or None,
                    "aliases_text": _text(record.get("别名")) or None,
                    "source_type": extracted.get("产品来源") or None,
                    "origin_area": _extract_origin(extracted.get("产品来源", "")),
                    "history_background": extracted.get("故事背景") or None,
                    "summary": extracted.get("总结") or None,
                    "suitable_level": _text(record.get("匹配人群")) or None,
                    "base_spec": _text(record.get("基础规格")) or None,
                    "base_price_text": _text(record.get("价格")) or None,
                    "raw_basic_info": raw_basic_info or None,
                }
                varieties.append(variety)
                for trait_type, trait_value in _extract_traits(raw_basic_info, extracted):
                    traits.append(
                        {
                            "variety_id": None,
                            "variety_name": variety_name,
                            "category_name": sheet_name,
                            "trait_type": trait_type,
                            "trait_value": trait_value,
                            "keywords": _keywords(trait_value),
                        }
                    )
                    chunks.append(
                        _chunk(
                            "orchid_variety_traits",
                            "variety",
                            variety_name,
                            sheet_name,
                            trait_type,
                            f"{variety_name} - {trait_type}",
                            trait_value,
                        )
                    )
                for value_type in VALUE_LABELS:
                    content = extracted.get(value_type, "")
                    if not content:
                        continue
                    value_points.append(
                        {
                            "variety_id": None,
                            "variety_name": variety_name,
                            "category_name": sheet_name,
                            "value_type": value_type,
                            "title": f"{variety_name}{value_type}",
                            "content": content,
                            "keywords": _keywords(content),
                        }
                    )
                    chunks.append(
                        _chunk(
                            "orchid_value_points",
                            "variety",
                            variety_name,
                            sheet_name,
                            value_type,
                            f"{variety_name} - {value_type}",
                            content,
                        )
                    )
        elif sheet_name == COMMON_KNOWLEDGE_SHEET:
            for record in _read_records(sheet, ["分类", "共享话术"]):
                category = _text(record.get("分类"))
                content = _text(record.get("共享话术"))
                if not category or not content:
                    continue
                row = {
                    "knowledge_category": category,
                    "knowledge_type": _knowledge_type(category),
                    "applies_to_category": _applies_to_category(category),
                    "content": content,
                }
                common_knowledge.append(row)
                chunks.append(
                    _chunk(
                        "orchid_common_knowledge",
                        "common_knowledge",
                        None,
                        row["applies_to_category"],
                        "通用知识",
                        category,
                        content,
                    )
                )
        elif sheet_name == SKU_SHEET:
            previous_name = ""
            for record in _read_records(sheet, ["品种", "品名"]):
                name = _text(record.get("品名")) or previous_name
                if not name:
                    continue
                previous_name = name
                price_text = _text(record.get("到手价"))
                skus.append(
                    {
                        "category_name": _text(record.get("品种")) or None,
                        "variety_name": name,
                        "seedling_count": _text(record.get("苗数")) or None,
                        "package_spec": _text(record.get("规格")) or None,
                        "flower_bud_status": _text(record.get("有无花苞")) or None,
                        "price": _price(price_text),
                        "price_text": price_text or None,
                    }
                )
        elif sheet_name == SALES_COPY_SHEET:
            for record in _read_records(sheet, ["填写人", "品种名"]):
                variety_name = _text(record.get("品种名"))
                if not variety_name:
                    continue
                row = {
                    "writer_name": _text(record.get("填写人")) or None,
                    "variety_name": variety_name,
                    "target_audience": _text(record.get("适合人群")) or None,
                    "product_background": _text(record.get("产品背景")) or None,
                    "leaf_posture": _text(record.get("叶姿")) or None,
                    "petal_type": _text(record.get("瓣型")) or None,
                    "flower_color": _text(record.get("花色")) or None,
                    "fragrance": _text(record.get("香味")) or None,
                    "flowering_period": _text(record.get("花期")) or None,
                    "care_difficulty": _text(record.get("养护难度")) or None,
                    "usage_scene": _text(record.get("使用场景")) or None,
                    "selling_points": _text(record.get("卖点")) or None,
                }
                sales_copy.append(row)
                for field, chunk_type in [
                    ("target_audience", "适合人群"),
                    ("product_background", "产品背景"),
                    ("leaf_posture", "叶姿"),
                    ("petal_type", "瓣型"),
                    ("flower_color", "花色"),
                    ("fragrance", "香味"),
                    ("flowering_period", "花期"),
                    ("care_difficulty", "养护难度"),
                    ("usage_scene", "使用场景"),
                    ("selling_points", "卖点"),
                ]:
                    content = row.get(field)
                    if content:
                        chunks.append(
                            _chunk(
                                "orchid_sales_copy",
                                "sales_copy",
                                variety_name,
                                None,
                                chunk_type,
                                f"{variety_name} - {chunk_type}话术",
                                content,
                            )
                        )
        elif sheet_name == HOT_BREAKDOWN_SHEET:
            for record in _read_records(sheet, ["品种名", "品类"]):
                variety_name = _clean_variety_name(record.get("品种名"))
                if not variety_name:
                    continue
                row = {
                    "variety_name": variety_name,
                    "category_name": _text(record.get("品类")) or None,
                    "status_history_supply_price_authenticity": _first_matching(record, "品种地位"),
                    "aesthetic_traits": _first_matching(record, "瓣"),
                    "cultivation_care": _first_matching(record, "栽培"),
                    "consensus_reputation": _first_matching(record, "共识"),
                }
                row["raw_text"] = "\n".join(value for value in row.values() if isinstance(value, str) and value)
                hot_breakdowns.append(row)
                for field, chunk_type in [
                    ("status_history_supply_price_authenticity", "专业拆解"),
                    ("aesthetic_traits", "产品特征"),
                    ("cultivation_care", "养护"),
                    ("consensus_reputation", "口碑共识"),
                ]:
                    content = row.get(field)
                    if content:
                        chunks.append(
                            _chunk(
                                "orchid_hot_breakdowns",
                                "hot_breakdown",
                                variety_name,
                                row["category_name"],
                                chunk_type,
                                f"{variety_name} - {chunk_type}",
                                content,
                            )
                        )

    return OrchidImportPayload(
        categories=categories,
        varieties=varieties,
        traits=traits,
        value_points=value_points,
        skus=skus,
        common_knowledge=common_knowledge,
        sales_copy=sales_copy,
        hot_breakdowns=hot_breakdowns,
        knowledge_chunks=chunks,
    )


def payload_to_json(payload: OrchidImportPayload, *, limit: int | None = None) -> str:
    data = payload.to_dict()
    if limit is not None:
        for key, value in list(data.items()):
            if isinstance(value, list):
                data[key] = value[:limit]
    return json.dumps(data, ensure_ascii=False, indent=2)


def _read_records(sheet, required_headers: list[str]) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = None
    headers: list[str] = []
    for index, row in enumerate(rows):
        candidate = [_text(cell) for cell in row]
        if all(header in candidate for header in required_headers):
            header_index = index
            headers = candidate
            break
    if header_index is None:
        return []

    records = []
    for row in rows[header_index + 1 :]:
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        if any(_text(value) for value in record.values()):
            records.append(record)
    return records


def _extract_basic_info(text: str) -> dict[str, str]:
    normalized = _normalize_text(text)
    if not normalized:
        return {}
    matches = list(
        re.finditer(
            r"(?:^|\n)\s*(?:\d+\s*[.、．]\s*)?([\u4e00-\u9fa5A-Za-z/（）()]{2,12})\s*[：:；;]",
            normalized,
        )
    )
    sections: dict[str, str] = {}
    for index, match in enumerate(matches):
        raw_label = match.group(1).strip()
        label = SECTION_LABELS.get(raw_label)
        if not label:
            continue
        end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        content = normalized[match.end() : end].strip()
        if content:
            sections[label] = _clean_section_content(content)
    if "总结" not in sections:
        summary_match = re.search(r"(?:^|\n)\s*总结[：:]\s*(.+)$", normalized)
        if summary_match:
            sections["总结"] = summary_match.group(1).strip()
    return sections


def _extract_traits(text: str, sections: dict[str, str]) -> list[tuple[str, str]]:
    found: dict[str, str] = {}
    source = "\n".join(value for value in [text, sections.get("产品特征", "")] if value)
    for label in sorted(TRAIT_LABELS, key=len, reverse=True):
        pattern = rf"{re.escape(label)}\s*[：:-]\s*([^\n。；;|]+(?:[。；;][^\n|：:]*)?)"
        match = re.search(pattern, source)
        if match:
            normalized = "香味" if label == "香气" else label
            found[normalized] = _clean_section_content(match.group(1))
    return [(key, value) for key, value in found.items() if value]


def _chunk(
    source_table: str,
    entity_type: str,
    variety_name: str | None,
    category_name: str | None,
    chunk_type: str,
    chunk_title: str,
    content: str,
) -> dict[str, Any]:
    return {
        "source_table": source_table,
        "source_id": None,
        "entity_type": entity_type,
        "variety_name": variety_name,
        "category_name": category_name,
        "chunk_type": chunk_type,
        "chunk_title": chunk_title,
        "content": content,
        "embedding_json": None,
    }


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").strip()


def _normalize_text(value: str) -> str:
    return _text(value).replace("|", "\n")


def _clean_section_content(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip(" \n\t")


def _extract_origin(source: str) -> str | None:
    match = re.search(r"(?:产自|产于|下山于|原产于)([^，。,；;\n]+)", source)
    return match.group(1).strip() if match else None


def _keywords(text: str) -> str | None:
    words = []
    for word in ["老种", "下山", "组培", "自然分株", "荷瓣", "梅瓣", "素心", "红花", "浓香", "勤花", "好养", "春化"]:
        if word in text:
            words.append(word)
    return ",".join(words) or None


def _knowledge_type(category: str) -> str:
    if "解释" in category or "种源" in category:
        return "术语解释"
    if "劣势" in category:
        return "品类劣势"
    if "优势" in category:
        return "品类优势"
    if "养" in category:
        return "养护知识"
    if "文化" in category:
        return "文化价值"
    return "销售话术"


def _applies_to_category(category: str) -> str | None:
    for name in VARIETY_SHEETS:
        if name != "其他品类" and name in category:
            return name
    return None


def _price(value: str) -> float | None:
    match = re.search(r"\d+(?:\.\d+)?", value)
    return float(match.group(0)) if match else None


def _clean_variety_name(value: Any) -> str:
    text = _text(value)
    for separator in ["|", "\n"]:
        if separator in text:
            text = text.split(separator, 1)[0]
    return text.strip()


def _first_matching(record: dict[str, Any], prefix: str) -> str | None:
    for key, value in record.items():
        if str(key).startswith(prefix):
            text = _text(value)
            return text or None
    return None
