from pathlib import Path

from openpyxl import load_workbook

from app.domains.sales.talk_script.repository import replace_talk_script_library


REQUIRED_SHEETS = {
    "scene_index",
    "question_cluster",
    "template_library",
    "field_dictionary",
}
SCENE_STATUSES = {"active", "disabled"}
GENERAL_STATUSES = {"active", "disabled", "need_review"}


class TalkScriptImportError(ValueError):
    pass


def import_talk_script_excel(path: str | Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    missing = REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing:
        raise TalkScriptImportError(f"Excel 缺少 sheet: {', '.join(sorted(missing))}")

    scenes = _read_sheet(workbook["scene_index"])
    questions = _read_sheet(workbook["question_cluster"])
    templates = _read_sheet(workbook["template_library"])
    _validate(scenes, questions, templates)
    replace_talk_script_library(scenes=scenes, questions=questions, templates=templates)
    return {
        "scene_count": len(scenes),
        "question_count": len(questions),
        "template_count": len(templates),
    }


def _read_sheet(sheet) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            record[header] = str(value).strip() if value is not None else ""
        if any(value for value in record.values()):
            records.append(record)
    return records


def _validate(scenes: list[dict], questions: list[dict], templates: list[dict]) -> None:
    errors: list[str] = []
    scene_ids = _unique_ids(scenes, "scene_id", errors)
    question_ids = _unique_ids(questions, "question_id", errors)
    template_ids = _unique_ids(templates, "template_id", errors)

    for row in scenes:
        status = row.get("status") or "active"
        if status not in SCENE_STATUSES:
            errors.append(f"scene_index.status 非法: {row.get('scene_id')}={status}")
        if not row.get("scene_name"):
            errors.append(f"scene_index.scene_name 不能为空: {row.get('scene_id')}")

    for row in questions:
        question_id = row.get("question_id")
        status = row.get("status") or "active"
        if status not in GENERAL_STATUSES:
            errors.append(f"question_cluster.status 非法: {question_id}={status}")
        if row.get("scene_id") not in scene_ids:
            errors.append(f"question_cluster.scene_id 不存在: {question_id}")
        if not row.get("default_template_id"):
            errors.append(f"question_cluster.default_template_id 不能为空: {question_id}")
        elif row.get("default_template_id") not in template_ids:
            errors.append(f"question_cluster.default_template_id 不存在: {question_id}")
        if not row.get("standard_question"):
            errors.append(f"question_cluster.standard_question 不能为空: {question_id}")

    active_template_by_question: set[str] = set()
    for row in templates:
        template_id = row.get("template_id")
        status = row.get("status") or "active"
        if status not in GENERAL_STATUSES:
            errors.append(f"template_library.status 非法: {template_id}={status}")
        if row.get("question_id") not in question_ids:
            errors.append(f"template_library.question_id 不存在: {template_id}")
        if not row.get("answer_default"):
            errors.append(f"template_library.answer_default 不能为空: {template_id}")
        if status == "active":
            active_template_by_question.add(row.get("question_id") or "")

    for row in questions:
        if (row.get("status") or "active") == "active":
            question_id = row.get("question_id") or ""
            if question_id not in active_template_by_question:
                errors.append(f"active question_id 缺少 active template: {question_id}")

    if errors:
        raise TalkScriptImportError("; ".join(errors))


def _unique_ids(rows: list[dict], field: str, errors: list[str]) -> set[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for row in rows:
        value = row.get(field) or ""
        if not value:
            errors.append(f"{field} 不能为空")
            continue
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    for value in sorted(duplicates):
        errors.append(f"{field} 重复: {value}")
    return seen
