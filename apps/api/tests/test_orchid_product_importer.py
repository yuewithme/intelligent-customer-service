from pathlib import Path

from openpyxl import Workbook

from app.domains.catalog.orchid_products.excel_importer import build_import_payload


def _save_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "建兰"
    ws.append(["产品经理-怡斌填写", None, None, None, None, None])
    ws.append(["品种名", "别名", "产品基础信息", "匹配人群", "基础规格", "价格"])
    ws.append(
        [
            "满堂红",
            "红满堂",
            "1.产品来源：下山兰\n2.故事背景：建兰八大红花之一。\n"
            "3.产品特征：花色：鲜红。瓣型：荷瓣。香味：浓香。花期：夏秋。\n"
            "4.差异性：寓意好，皮实好养。\n5.市场价：30-100元/盆",
            "L1-L2",
            "3-5苗",
            "68-88",
        ]
    )

    ws = wb.create_sheet("通用知识点")
    ws.append(["说明书", None])
    ws.append(["分类", "共享话术"])
    ws.append(["兰花种源解释", "老种：经过长时间自然环境驯化的自然苗"])

    ws = wb.create_sheet("链接详情")
    ws.append(["品种", "品名", "苗数", "规格", "有无花苞", "到手价"])
    ws.append(["建兰", "富贵金龙", "3-5苗", "裸苗", "带花剑", 68])
    ws.append(["建兰", None, "3-5苗", "国风盆", None, 88])

    ws = wb.create_sheet("私域运营团队2")
    ws.append(["说明书", None])
    ws.append(
        [
            "填写人",
            "品种名",
            "适合人群",
            "产品背景",
            "叶姿",
            "瓣型",
            "花色",
            "香味",
            "花期",
            "养护难度",
            "使用场景",
            "卖点",
        ]
    )
    ws.append(
        [
            "振山",
            "满堂红",
            "新手兰友",
            "经典红花品种",
            "叶子挺拔",
            "荷瓣",
            "红花",
            "浓香",
            "夏秋开花",
            "好养",
            "客厅、办公室",
            "好看好养寓意好",
        ]
    )

    ws = wb.create_sheet("热门品种拆解")
    ws.append(["说明", None, None, None, None, None])
    ws.append(
        [
            "品种名",
            "品类",
            "品种地位与历史\n存量与供需\n价格体系\n真假鉴别",
            "瓣（最重要的审美标准）\n色（颜色与质感）\n香（香型即灵魂）\n姿（叶姿花姿态与整体）\n神（花守与开品）",
            "栽培与养护",
            "共识\n口碑",
        ]
    )
    ws.append(
        [
            "宋梅\n春兰四大天王之首",
            "春兰",
            "四大天王之首",
            "梅瓣，幽香",
            "需要春化",
            "经典名品",
        ]
    )

    wb.save(path)


def test_build_import_payload_classifies_excel_sheets(tmp_path):
    path = tmp_path / "orchid.xlsx"
    _save_workbook(path)

    payload = build_import_payload(path)

    assert payload.counts == {
        "categories": 1,
        "varieties": 1,
        "traits": 4,
        "value_points": 4,
        "skus": 2,
        "common_knowledge": 1,
        "sales_copy": 1,
        "hot_breakdowns": 1,
        "knowledge_chunks": 23,
    }
    assert payload.categories[0]["category_name"] == "建兰"
    assert payload.varieties[0]["variety_name"] == "满堂红"
    assert payload.varieties[0]["source_type"] == "下山兰"
    assert payload.varieties[0]["history_background"] == "建兰八大红花之一。"
    assert payload.hot_breakdowns[0]["variety_name"] == "宋梅"


def test_sku_rows_inherit_previous_variety_name(tmp_path):
    path = tmp_path / "orchid.xlsx"
    _save_workbook(path)

    payload = build_import_payload(path)

    assert [row["variety_name"] for row in payload.skus] == ["富贵金龙", "富贵金龙"]
    assert payload.skus[0]["price"] == 68.0
    assert payload.skus[1]["package_spec"] == "国风盆"


def test_sales_copy_is_kept_as_answer_material_not_master_data(tmp_path):
    path = tmp_path / "orchid.xlsx"
    _save_workbook(path)

    payload = build_import_payload(path)

    sales_copy = payload.sales_copy[0]
    assert sales_copy["writer_name"] == "振山"
    assert sales_copy["target_audience"] == "新手兰友"
    assert sales_copy["selling_points"] == "好看好养寓意好"
    assert payload.varieties[0]["history_background"] == "建兰八大红花之一。"
